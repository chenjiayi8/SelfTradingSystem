# -*- coding: utf-8 -*-
"""
Created on Wed Mar  4 10:35:00 2020

@author: Frank
"""


#import re
#import pickle
#import os
#resultFile = os.path.join(os.getcwd(), 'df_lastTasks_traded.out')
#f = open(resultFile, 'rb')
#tasks_valided = pickle.load(f)
#f.close()
import pandas as pd
import openpyxl as op
from datetime import time
from collections import OrderedDict
from xlwings.constants import AutoFillType
from SelfTradingSystem.util.convert import getTodayDate, dateTimeToDateStr

def getTargetArea(sht, startCol, endCol=[], startRow=0, endRow=0):
    if len(endCol) > 0:
        header = sht.range(startCol+str(1), endCol+str(1)).value
    else:
        header = sht.range(startCol+str(1)).value
    numLastRow_Region = sht.range(startCol+str(1)).current_region.last_cell.row
    for i in range(numLastRow_Region+1):
        cellValue = sht.range(startCol+str(i+1)).value
        if cellValue == None or cellValue == "None" or cellValue == "nan":
            numLastRow = i
            break
    
    if startRow == 0:
        startRow = 2
    elif startRow < 0:
        startRow += numLastRow + 1
    
    if endRow == 0:
        endRow = numLastRow
    elif endRow < 0:
        endRow += numLastRow + 1
      
    if len(endCol) > 0:    
        data =  sht.range(startCol+str(startRow), endCol+str(endRow)).value
        if numLastRow == 2:
            data = [data]
        return pd.DataFrame(data=data, columns=header)
    else:
        data =  sht.range(startCol+str(2), startCol+str(endRow)).value
        return pd.Series(data=data, name=header).to_frame()

def sheetToDF(sht):
    return sht.range('A1').options(pd.DataFrame, 
                         header=1,
                         index=False, 
                         expand='table').value

def indCell(colStr, rowNumber):
    if type(colStr) is str:
        return colStr+str(rowNumber)
    else:
        return chr(ord('A') + colStr -1)+str(rowNumber)

def getColumnStr(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def codeToStr(inputNumber):
    type_input = type(inputNumber)
    if type_input is str:
        return inputNumber
    if type_input is float:
        numberStr = ""
        divisor = 100000
        for i in range(6):
            numberStr += str(int(inputNumber//divisor))
            inputNumber = inputNumber%divisor
            divisor /= 10
        return numberStr

shtStrDic = {"自动二八":"自动二八",
          "行业轮动":"行业轮动",
          "日线交易":"日线交易",
          "目标市值两融":"目标市值两融",
          "目标市值":"目标市值",
          "两融网格":"两融网格",
          "简易网格":"简易网格",
          "普通网格":"普通网格",
          "趋势卖":"趋势份额",
          "趋势买":"趋势份额"
          }
sht_军工普通_Dic = {"军工小网格":"军工小网格",
              "军工无底仓":"军工网格无底仓",
              "军工PB大网格":"军工大网格"
              }

sht_config_Dic = {"军工小网格":["A33", "L33", [2,2,1,1,1,2,2,2,2,2,2,2]], #1 is filled with input, 2 is filled with format or previous row
                  "券商网格":["A30", "L30", [2,2,1,1,1,2,2,2,2,2,2,2]],
                  "自动网格":["W1", "AF1", [1,1,1,1,1,2,3,2,2,2]],
                  "目标市值":["AX1", "BK1", [1,1,1,1,1,2,3,2,2,2,0,0,0,2]],
                  "目标市值两融":["AX1", "BM1", [1,1,1,1,1,2,3,2,0,2,2,2,0,0,0,2]],
                  "自动二八":["AH1", "AS1", [1,1,1,1,1,2,2,3,2,2,2,2]],
                  "军工网格无底仓":["A33", "L33", [2,2,1,1,1,2,2,2,2,2,2,2]],
                  "军工大网格":["A37", "M37", [2,2,1,1,1,2,2,0,2,2,2,2,2]],
                  "油气网格":["A30", "L30", [2,2,1,1,1,2,2,2,2,2,2,2]],
                  "两融网格":["AR1", "BH1", [1,1,1,1,1,2,3,2,2,2,2,2,0,2,2,2,2]],
                  "简易网格":["AT1", "BF1", [1,1,1,1,1,2,3,2,2,2,2,2,2]],
                  "行业轮动":["AE1", "AO1", [1,1,1,1,1,2,3,2,2,3,2]],
                  "日线交易":["AG1", "AR1", [1,1,1,1,1,2,2,3,2,2,2,2]],
                  "趋势份额":["AD1", "AQ1", [1,1,1,1,1,1,1,2,3,2,2,2,2,2]],
                  }
letters = list(map(chr, range(ord('A'), ord('Z')+1)))
colStr_Dic = {}
colNum_Dic = {}
for i in range(1, 1000):
    colStr = ''.join(getColumnStr(i))
    colStr_Dic[i] = colStr
    colNum_Dic[colStr] = i

def seperateInd(ind):
    numStr = []
    colStr = []
    for r in ind:
        try:
            int(r)
            numStr.append(r)
        except:
            colStr.append(r)
            pass
    rowNum = int(''.join(numStr))
    colStr = ''.join(colStr)
    colNum  = colNum_Dic[colStr]
    return colStr, colNum, rowNum

def getInputFromTradedTask(df_row):
    name   = df_row.loc['Name']
    code   = codeToStr(df_row.loc['Code'])
    time   = getTodayDate()
    amount = df_row.loc['Amount']
    price  = df_row.loc['成交价']
    remark = df_row.loc['Remark']
    if remark == "趋势买":
        return [name, code, time, amount*-1,price, round(price*amount*-1, 2), remark]
    elif remark == "趋势卖":
        return [name, code, time, amount*-1,price, 0, remark]
    else:
        return [name, code, time, amount, price, remark]

def getInputFromValidTask(df_row):
    name   = df_row.loc['基金名称记录']
    code   = codeToStr(df_row.loc['基金代码记录'])
    time   = getTodayDate()
    amount = df_row.loc['锁卖份额']
    money  = df_row.loc['锁买金额']
    price  = df_row.loc['操作价格']
    remark = df_row.loc['策略名称']
    return [name, code, time, amount,price, money,remark]

def getTargetSheetStr(sysObj, input_list, isMomentumShare=False):
    remark = input_list[-1]
    sheetName = shtStrDic[remark]
    if "普通网格" == sheetName:
        if "军工" in input_list[0]:
            sheetName = sht_军工普通_Dic[input_list[0]]
        if "油气" in input_list[0]:
            sheetName = "油气网格"
        if "证券" in input_list[0]:
            sheetName = ["券商网格", "自动网格"]
    if isMomentumShare:
        if type(sheetName) is str:
            sheetName = [sheetName, "趋势份额"]
        else:
            sheetName.append("趋势份额")
    return sheetName

def checkSheetNameInBook(sheetName, sheetNameList):
    if type(sheetName) is str:
        return sheetName in sheetNameList
    else:
        flag = True
        for name in sheetName:
            if name not in sheetNameList:
                flag = False
        return flag

def moveFormulaForOneRow(oldFormula):
    oldFormulaList = [ c for c in oldFormula]
    numLocation = []
    for i in range(len(oldFormulaList)):
        try: 
           int(oldFormulaList[i])
           numLocation.append(i)
        except:
            pass
    numIdxs = []
    for i in range(len(numLocation)):
        numIdx = [numLocation[i]]
        if not any(numLocation[i] in L for L in numIdxs):
            if numLocation[i] != numLocation[-1]:
                j = i
                
                while j < len(numLocation) - 1:
                    if numLocation[j] + 1 == numLocation[j+1]:
                        numIdx.append(numLocation[j+1])
                        j += 1
                    else:
                        break
            numIdxs.append(numIdx)
    for i in range(len(numIdxs)):
        numIdx = numIdxs[i]
        numStr = [ oldFormulaList[j] for j in numIdx]
        charBeforeNumStr = oldFormulaList[numIdx[0]-1]
        if charBeforeNumStr in letters:
            num = int(''.join(numStr))
            num_new = num + 1
            num_new_str = str(num_new)
            num_new_str = [c for c in num_new_str]
            if len(num_new_str) > len(numStr):
                count = 0 
                for j in numIdx:
                    oldFormulaList[j] = num_new_str[count]
                    count += 1
                oldFormulaList.insert(numIdx[-1]+1, num_new_str[-1]) # normally only increase by one
                for j in range(i+1, len(numIdxs)):
                    temp_numIdx = numIdxs[j]
                    temp_numIdx = [c+1 for c in temp_numIdx]
                    numIdxs[j] = temp_numIdx
            else:
                count = 0 
                for j in numIdx:
                    oldFormulaList[j] = num_new_str[count]
                    count += 1
        
    return ''.join(oldFormulaList)
#    lastRowNumStr = str(lastRowNum)
#        cell_formula_new  = cell_formula.replace(str(numLastRow), str(numLastRow+1))
#        lastRowIdxs = [m.start() for m in re.finditer(str(numLastRow-1),cell_formula_new)]
#        subStrLength = len(str(numLastRow-1))
#        cell_formula_new  = cell_formula_new.replace(str(numLastRow-1), str(numLastRow))


def sheetWriter(sysObj, sheetStr, input_list, isMomentumShare=False):
    sht = sysObj.wb.sheets[sheetStr]
    sht_config = sht_config_Dic[sheetStr]
    colStr_start, colNum_start, rowNum_start = seperateInd(sht_config[0])
    colStr_end, colNum_end, rowNum_end = seperateInd(sht_config[1])
    numLastRow_Region = sht.range(sht_config[0]).current_region.last_cell.row
    for i in range(rowNum_start, numLastRow_Region+1):
        cellValue = sht.range(colStr_start+str(i+1)).value
        if cellValue == None or cellValue == "None" or cellValue == "nan":
            numLastRow = i
            break
    sourceIdxList =  [colStr_Dic[c]+str(numLastRow) for c in range(colNum_start,colNum_end+1)]  
    targetIdxList =  [colStr_Dic[c]+str(numLastRow+1) for c in range(colNum_start,colNum_end+1)]
#    sourceValueList   = sht.range(sourceIdxList[0], sourceIdxList[-1]).value
    sourceFormulaList = list(sht.range(sourceIdxList[0], sourceIdxList[-1]).formula[0])
    config = sht_config[-1]
    if len(config) != len(sourceFormulaList):
        raise ("len of config {} does not match len of sourceValueList {}".format(len(config), len(sourceFormulaList)))
    for i in range(len(sourceFormulaList)):
        config_value = config[i]
        source_idx   = sourceIdxList[i]
        target_idx   = targetIdxList[i]
        # cell_formula = sourceFormulaList[i]
        # cell_formula_new = moveFormulaForOneRow(cell_formula)
        if config_value > 1:
            sht.range(source_idx).api.AutoFill(sht.range(source_idx+':'+target_idx).api, AutoFillType.xlFillDefault)
            pass #autofill
            # if config_value == 2:
            #     sht.range(target_idx).formula = cell_formula_new
            # else:
            #     sht.range(target_idx).formula_array = cell_formula_new
        elif config_value == 1:
            sht.range(target_idx).value = input_list[i]
    if isMomentumShare:
        sht.range(colStr_Dic[colNum_end+1]+str(numLastRow+1)).value = "趋势份额"

def excelWriter(sysObj, df_row, isMomentumShare=False):
    if not isMomentumShare:
        input_list = getInputFromTradedTask(df_row)
    else:
        input_list = getInputFromValidTask(df_row)
    sheetStrs = getTargetSheetStr(sysObj, input_list, isMomentumShare)
    if type(sheetStrs) is str:
        sheetWriter(sysObj, sheetStrs, input_list, isMomentumShare)
    else:
        for sheetStr in sheetStrs:
            sheetWriter(sysObj, sheetStr, input_list, isMomentumShare)
    
  
def updateTradedTasks(sysObj, df_lastTasks_traded):
#    df_lastTasks_traded = df_lastTasks_traded.sort_values(by=['Amount'])
    for i in range(len(df_lastTasks_traded)):
        df_row = df_lastTasks_traded.iloc[i].copy()
        excelWriter(sysObj, df_row)

def writeMomentTasks(sysObj, tasks_valided):
    for i in range(len(tasks_valided)):
        df_row = tasks_valided.iloc[i].copy()
        excelWriter(sysObj, df_row, True)
    


def toStringWithFormat(cell):
    value = cell.value
    number_format = cell.number_format
    data_type = cell.data_type
    if number_format == None or value == None:
        value = ''
    elif type(value) is bool:
        value = str(value)
    elif cell.is_date:
        if type(value) is not time:
            value = dateTimeToDateStr(value)
        else:
            value = value.strftime("%H:%M:%S")
    elif data_type in ['s', 'e'] :
        pass
    elif number_format in ['0.00', '0.0', '0.00000', '0.000000']:
        value = '{0:.2f}'.format(value)
    elif number_format in ['0.00%', '0.0%', '0%', '0.000%']:
        value = '{0:.2%}'.format(value)
    elif  number_format in ['0.0000']:
        value = '{0:.4f}'.format(value)
    elif  number_format in ['0.000']:
        value = '{0:.3f}'.format(value)
    elif  number_format in ['0', '@']:
        if type(value) == float:
             value = '{0:.2f}'.format(value)
        else:
            value = '{:d}'.format(value)
    elif number_format == 'General':
        pass
    else:
        print(number_format)
    return value
        
def dfToDatabaseDF(df, columns=None):
    if columns is None:
        col_num = list(df.columns)
        col_str = [colStr_Dic[i+1] for i in col_num]
    else:
        col_str = list(columns)
        
    df_new = pd.DataFrame(columns=col_str)
    target_columns = col_str[:len(df.columns)]
    df_new = df_new.append(dict(zip(target_columns, df.columns)), ignore_index=True)
    for rowID in df.index:
       df_new = df_new.append(dict(zip(target_columns, list(df.loc[rowID, :]))), ignore_index=True) 

    return df_new
    


def excelToDFs(xlsxPath):
    # wb = op.load_workbook(xlsxPath, read_only=True, data_only=True)
    wb = op.load_workbook(xlsxPath, data_only=True)
    print("Finishing reading {}".format(xlsxPath))
    sheetnames = wb.sheetnames
    dfs = []
    for sheetname in sheetnames:
        ws = wb[sheetname]
        data = []
        for r in range(1, ws.max_row+1):
            row_data = []
            for c in range(1, ws.max_column+1):
                cell = ws.cell(r, c)
                row_data.append(toStringWithFormat(cell))
            data.append(row_data)
            
        df = pd.DataFrame.from_records(data)
        col_num = list(df.columns)
        col_str = [colStr_Dic[i+1] for i in col_num]
        df.columns = col_str        
        dfs.append(df)
    wb.close()
    return OrderedDict(zip(sheetnames, dfs))


def removeMargin(df):
    columns = list(df.columns)
    columns_drop = []
    for col in columns:
        col_list = df[col].to_list()
        cell_empty = [len(cell)==0 for cell in col_list]
        if all(cell_empty):
            columns_drop.append(col)
    rows = list(range(len(df)))
    rows_drop = []
    for row in rows:
        row_list = df.iloc[row].to_list()
        cell_empty = [len(cell)==0 for cell in row_list]
        if all(cell_empty):
            rows_drop.append(row)
    df = df.drop(columns=columns_drop, index=rows_drop).reset_index(drop=True)
    return df

if __name__ == '__main__':
    # xlsxPath = '本金账本.xlsx'
    # dfs_dict = excelToDFs(xlsxPath)
    from SelfTradingSystem.io.database import Database
    from SelfTradingSystem.core.trade import Trade
    db_path = 'Resources.db'
    sql = Database(db_path)
    # sql.createDB(xlsx_path, db_path)
    # print(sql.getLastRows('S000985', 10))
    # sleep(5)

    pass
